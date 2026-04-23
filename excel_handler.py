"""
File: excel_handler.py
Purpose: Handles all Excel read and write operations for Venire Automation
"""

# Standard Library Imports
import os

# Third-Party Library Imports
import openpyxl # 3.1.5

# CONSTANTS
NAME_COLUMN = 3
DOB_COLUMN = 2
JUROR_COLUMN = 1
OUTCOME_COLUMN = 4

# PARSE CONSTANTS
MAX_SPLIT_PARAMETER = 1
FIRST_NAME_SPLIT_PARAMETER = 0
COMMA_DELIMITER = ","

# FUNCTIONS
def load_workbook_safe(file_path: str):
    """
    Opens and returns an openpyxl workbook object.
    Raises FileNotFoundError if the file does not exist.

    Args:
        file_path: Path to the Excel file ex: 'venire.xlsx'

    Returns:
        openpyxl Workbook object
    """
    if not os.path.exists(file_path): # If the file path does not exist is not there.
        raise FileNotFoundError(f"File not found: {file_path}") # Raise an error if the Excel sheet cannot be found.
    return openpyxl.load_workbook(file_path) # Return the open Excel sheet if it is found.

def parse(name: str)-> tuple:
    """
    Splits a name string formatted as 'Last, First Middle'
    into first and last name components.

    Args:
        name: Full name string ex: 'Smith, John Michael'

    Returns:
        (first_name, last_name) as strings ex: ('John', 'Smith')
        (None, None) if name is empty or missing a comma
    """
    if not name or "," not in name: # Checks to if the name is empty or does not have a comma
        return None, None # If empty or does not have a comma return None, None

    left, right = name.split(COMMA_DELIMITER, MAX_SPLIT_PARAMETER) # Split the name by the comma and at the first comma it sees.

    last_name = left.strip() # Removes whitespace
    right = right.strip() # Removes whitespace with the first
    if not right:
        return None, None

    first_name = right.split()[FIRST_NAME_SPLIT_PARAMETER] # This takes the first half of the two names.

    return first_name, last_name

def format_dob(dob) -> str:
    """
    Formats a date of birth into mm/dd/yyyy string.
    Handles both datetime objects and plain strings.

    Args:
        dob: Date of birth as datetime object or string

    Returns:
        Formatted date string ex: '04/23/1985'
        None if the format is unrecognized
    """
    # Case 1 - datetime object
    if hasattr(dob, 'strftime'): # If dob has the strftime attribute
        return dob.strftime('%m/%d/%Y') # Convert it to a string

    # Case 2 - plain string
    if isinstance(dob, str): # If dob is just a plain string
        return dob.strip() # Keep it how it is and clean it up.

    return None

def build_juror_from_row(sheet, row: int) -> tuple:
    """
    Reads one row from the sheet and builds a juror dictionary.
    Calls parse() and format_dob() to clean the data.

    Args:
        sheet: openpyxl worksheet object
        row:   Row number to read from

    Returns:
        (juror_id, juror_data) as a tuple
        juror_id  : string ex: "1042"
        juror_data: dict with row, dob, first_name, last_name
        None if any cell is empty or parsing fails
    """
    juror_id = sheet.cell(row=row, column=JUROR_COLUMN).value # Grab the value that shows the Juror ID
    dob      = sheet.cell(row=row, column=DOB_COLUMN).value # Grabs the value that shows the DOB
    name     = sheet.cell(row=row, column=NAME_COLUMN).value # Grabs the value that shows the name

    # If any cell is empty skip this row
    if not juror_id or not dob or not name:
        return None

    first_name, last_name = parse(name) # Parse the name to only have first name and last name
    formatted_dob = format_dob(dob) # Format the dob to make sure that it is a string and can but used in CCIS

    if not first_name or not last_name or not formatted_dob:
        return None

    juror_data = {
        "row"        : row,
        "dob"        : formatted_dob,
        "first_name" : first_name,
        "last_name"  : last_name,
    }

    return str(juror_id), juror_data # Creates the dictionary and uses the juror ID as the key and the juror data as the value of the dictionary

def parse_juror_sheet(sheet) -> dict:
    """
    Loops through every row in the sheet and builds
    a dictionary of all valid jurors.

    Calls build_juror_from_row() for each row.
    Skips any row that returns None.

    Args:
        sheet: openpyxl worksheet object

    Returns:
        Dictionary of all valid jurors keyed by juror ID
        {
            "1042": {"row": 2, "dob": "04/23/1985", "first_name": "John", "last_name": "Smith"},
            "1043": {"row": 3, "dob": "06/15/1990", "first_name": "Jane", "last_name": "Doe"},
        }
        Empty dict if no valid jurors found
    """
    jurors = {}

    for row in range(1, sheet.max_row + 1):  # start at top of the list, max data knows where the data ends on the list
        result = build_juror_from_row(sheet, row)
        if result is None:  # skip empty or bad rows
            continue
        juror_id, juror_data = result # unpack the two things returned
        jurors[juror_id] = juror_data  # add to dictionary

    return jurors
