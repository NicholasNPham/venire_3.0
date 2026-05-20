# Report Builder

# THIRD-PARTY IMPORTS
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
from pypdf import PdfWriter, PdfReader

# STANDARD LIBRARY IMPORTS
import os
from pathlib import Path
import logging

# CONSTANT
ALL_PDF_FILENAME_EXTENSION = "*.pdf"
VENIRE_EXCEL_PDF_FILENAME = "VENIRE_EXCEL_REPORT.pdf"
VENIRE_FINAL_REPORT = "VENIRE_FINAL_REPORT.pdf"

#FUNCTION
def excel_to_pdf(excel_path: str, output_pdf_path: str) -> str:
    """
    Converts an Excel file to a PDF table.

    Args:
        excel_path: Path to the .xlsx file.
        output_pdf_path: Where to save the resulting PDF.

    Returns:
        The output_pdf_path on success.

    Example:
        excel_to_pdf("results/jurors.xlsx", "results/jurors_report.pdf")
    """
    if not os.path.exists(excel_path): # If the file path does not exist is not there.
        raise FileNotFoundError(f"File not found: {excel_path}") # Raise an error if the Excel sheet cannot be found.
    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else "" for cell in row])

    workbook.close()

    document = SimpleDocTemplate(output_pdf_path, pagesize=letter)

    document.build([Table(data)])

    return output_pdf_path

def merge_pdfs(pdf_paths: list[str], output_path: str) -> str:
    """
    Merges a list of PDFs into a single PDF file.

    Args:
        pdf_paths: Ordered list of PDF file paths to merge.
        output_path: Destination path for the merged PDF.

    Returns:
        The output_path on success.

    Example:
        merge_pdfs(["report.pdf", "juror_1.pdf", "juror_2.pdf"], "final.pdf")
    """
    writer = PdfWriter()

    for path in pdf_paths:
        reader = PdfReader(path)

        for page in reader.pages:
            writer.add_page(page)

    with open(output_path, "wb") as file:
        writer.write(file)

    return output_path

def combine(results_folder: str, excel_path: str, log: logging.Logger) -> None:
    """
    Merges all juror PDFs and the Excel report into a single combined PDF file.

    Args:
        results_folder: Folder containing all juror PDFs.
        excel_path: Path to the completed Excel sheet.
        log: Logger instance for recording progress and errors.

    Returns:
        None

    Example:
        combine("screenshots/2025-04-23", "jurors.xlsx", log)
    """
    results_path = Path(results_folder)

    if not results_path.exists():
        raise FileNotFoundError(f"Results folder not found: {results_folder}")

    juror_pdfs = sorted(results_path.glob(ALL_PDF_FILENAME_EXTENSION))

    if not juror_pdfs:
        log.error("No PDF files found in the results folder.")
        return

    excel_pdf_path = results_path / VENIRE_EXCEL_PDF_FILENAME
    excel_to_pdf(excel_path, str(excel_pdf_path))

    final_pdf_list = [str(excel_pdf_path)] + [str(pdf) for pdf in juror_pdfs]

    root_path = results_path.parent.parent
    output_path = root_path / VENIRE_FINAL_REPORT
    merge_pdfs(final_pdf_list, str(output_path))

    if excel_pdf_path.exists():
        excel_pdf_path.unlink()

    log.info(f"Combined PDF created at: {output_path}")

def prompt_and_combine(results_folder: str, excel_path: str, log: logging.Logger) -> None:
    """
    Prompts the user for confirmation before combining all juror PDFs and the Excel report into one file.

    Args:
        results_folder: Folder containing all juror PDFs.
        excel_path: Path to the completed Excel sheet.
        log: Logger instance for recording progress and errors.

    Returns:
        None

    Example:
        prompt_and_combine("screenshots/2025-04-23", "jurors.xlsx", log)
    """
    answer = input("Do you wish to combine the Excel report and all juror PDFs into one? (y/n): ").strip().lower()

    if answer != "y":
        log.info("Left alone — no changes made.")
        return

    combine(results_folder, excel_path, log)

